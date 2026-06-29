from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def format_excel(ws):

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=11
    )

    body_font = Font(size=10)

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Header formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # Body formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = center
            cell.border = thin

    # Auto column width
    for column in ws.columns:
        max_length = 0

        for cell in column:
            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except:
                pass

        ws.column_dimensions[
            get_column_letter(column[0].column)
        ].width = max_length + 3

    ws.freeze_panes = "A2"